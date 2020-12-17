__author__ = 'zn'
from openpyxl import load_workbook
import random
import datetime
import logging
from Common import myLogger2
class DoExcel:
    def __init__(self,excelpath):
        #初始化对象
        self.wb=load_workbook(excelpath)
        self.case_datas = self.wb["case_datas"]
        self.prepare_data=self.wb["prepare_data"]

     #从excel当中读取初始化数据，并对手机号码进行处理。

    def get_init_datas(self):
        init_datas = {}
        for index in range(2, self.prepare_data.max_row + 1):
            key = self.prepare_data.cell(row=index,column=1).value
            init_datas[key] = str(self.prepare_data.cell(row=index,column=2).value)

        init_datas["${noRecommend_phone}"] = str(int(init_datas["${init_phone}"]) + 1)
        init_datas["${phone}"] = str(int(init_datas["${init_phone}"]) + 2)
        #logging.info("初始化注册手机号数据为：{0}".format(init_datas))
        #{"${init_phone}":"18612340000","${phone}":"18612340001","${phone}":"18612340002"}
        return init_datas

    #for循环读取所有测试数据
    def getAllCaseDatas(self):
        All_caseDatas=[]
        for index in range(2,self.case_datas.max_row+1):
            case_data={}
            case_data["case_id"]=self.case_datas.cell(row=index,column=1).value
            case_data["api_name"]=self.case_datas.cell(row=index,column=5).value
            case_data["ip"]=self.case_datas.cell(row=index,column=6).value
            case_data["port"]=self.case_datas.cell(row=index,column=7).value
            case_data["interface"]=self.case_datas.cell(row=index,column=8).value
            case_data["method"]=self.case_datas.cell(row=index,column=9).value
            temp_request_data = self.case_datas.cell(row=index, column=10).value
            if self.case_datas.cell(row=index,column=12).value is not None:
               temp_sql_data=self.case_datas.cell(row=index,column=12).value
               #logging.info(temp_sql_data)
            #获取注册登录手机号初始值
            init_datas=self.get_init_datas()
            #random当前时间和字母获取流水号 通过循环替换请求数据里的serial流水号的值
            randomData={}
            randomDate={}
            for i in range(0,self.case_datas.max_row+1):
                 seed ="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ"
                 sa =[]
                 for i in range(0,2):
                     sa.append(random.choice(seed))
                     salt = ''.join(sa)
                 nowTime=datetime.datetime.now().strftime("%Y%m%d%H%M%S")
                 randomNum=random.randint(0,1000)
                 if randomNum<=1000:
                      randomNum=str(0)+str(randomNum)
                      uniqueNum=str(nowTime)+str(randomNum)+str(salt)
                      #组合产品开放日卖出时 当开放日修改为当前日期时自动替换请求数据中的"apdt"和"workdate"值为当前日期
                      nowDate=datetime.datetime.now().strftime("%Y%m%d")

                      if uniqueNum and nowDate not in randomData.values():
                          randomData["temp_serial"]=uniqueNum
                          randomDate["nowDate"]=nowDate

                 #替换请求数据和sql语句里的serial流水号的值
                 if temp_request_data is not None and temp_sql_data is not None and randomData is not None and randomDate is not None and len(init_datas) > 0:
                     #替换request data和sql语句里的serial值
                     for key,value in randomData.items():
                         if temp_request_data.find(key) != -1:
                            temp_request_data = temp_request_data.replace(key,value)
                            if temp_sql_data.find(key) !=-1:
                               temp_sql_data = temp_sql_data.replace(key,value)
                     #组合产品开放日卖出时 当开放日修改为当前日期时自动替换请求数据中的"apdt"和"workdate"值为当前日期
                     for key,value in randomDate.items():
                         if temp_request_data.find(key) != -1 :
                             temp_request_data = temp_request_data.replace(key,value)
                     #替换注册,登录模块里的手机号的值
                     for key,value in init_datas.items():
                         if temp_request_data.find(key) != -1:
                             temp_request_data = temp_request_data.replace(key,value)
                             if temp_sql_data.find(key) !=-1 and temp_sql_data is not None:
                                temp_sql_data = temp_sql_data.replace(key,value)

            logging.info("初始化之后的请求数据为：\n{0}".format(temp_request_data))
            logging.info("初始化之后的sql语句为：\n{0}".format(temp_sql_data))
            case_data["request_data"] = temp_request_data
            case_data["execute_sql"] = temp_sql_data
            case_data["expected_data"] = self.case_datas.cell(row=index,column=11).value
            case_data["affected_caseId"] = self.case_datas.cell(row=index,column=14).value
            case_data["compare_type"]=self.case_datas.cell(row=index,column=13).value
            case_data["wait_time"]=self.case_datas.cell(row=index,column=15).value
            #调用函数更新后续关联接口测试数据 requestdata和sql语句中的serialno
            if case_data["affected_caseId"] is not None:
                self.update_affectedApi_requestData(case_data["affected_caseId"],eval(case_data["request_data"])["serialno"])

            #判断本用例是否需要对响应结果进行解析
            # if self.case_datas.cell(row=index,column=14).value != None:
            #     case_data["related_exp"] = self.case_datas.cell(row=index,column=14).value
            All_caseDatas.append(case_data)
        return All_caseDatas

    #更新后续关联接口中的请求数据和sql查询语句
    #交易后续接口请求数据中需要前一个接口中的同一个serinal no来发起请求，首先获取前一个接口中的serialno  然后替换后续接口中的serialno(后续接口中的serinalno先用占位符表示)，sql同理
    def update_affectedApi_requestData(self,affected_caseId,AffectedData):
        for row_num in range(2,self.case_datas.max_row+1):
                if int(self.case_datas.cell(row=row_num,column=1).value) == int(affected_caseId) :
                     origin_requestData=self.case_datas.cell(row=row_num,column=10).value
                     origin_selectSql=self.case_datas.cell(row=row_num,column=12).value
                     self.case_datas.cell(row=row_num,column=10).value=str(origin_requestData).replace("${}",str(AffectedData))
                     self.case_datas.cell(row=row_num,column=12).value=str(origin_selectSql).replace("${}",str(AffectedData))

    #更新手机号初始化数据
    def update_init_data(self):
        init_data = self.prepare_data.cell(row=2,column=2).value
        self.prepare_data.cell(row=2, column=2).value = str(int(init_data) + 3)



    #修改excel后保存
    def save_excel(self,excelpath):
        try:
            self.wb.save(excelpath)
        except Exception as e:
            logging.info(e)








